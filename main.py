import xlrd
import xlwt
import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import date

today = date.today()
today1 = today.strftime("%d/%m/%Y")

def calculate_fuel_economy(kilometresDriven, litresUsed): # Function to calculate fuel economy by litres per 100km
    fuelEconomy =  (litresUsed / kilometresDriven) * 100
    return fuelEconomy

def insert_new_data(d,f):
    book = load_workbook('testDocument.xlsx') # Load workbook
    sheet = book.active # Load active sheet
    new_data = (today1, int(d), int(f))
    sheet.append(new_data)
    book.save('testDocument.xlsx')
    print("Data successfully added to spreadsheet")

print("Welcome to the Fuel Economy Calculator") # Welcome message
print("=======================================")
print("what would you like to do?")
print("1. Calculate fuel economy")
print("2. Get Summary Data From Spreadsheet")
type = input("Please enter 1 or 2: ") # User input to choose between 1 or 2
if type == "1":
    kilometresDriven = input("Please enter the number of kilometres driven: ") # User input for kilometres driven
    litresUsed = input("Please enter the number of litres used: ") # User input for litres used
    fuelEconomy = calculate_fuel_economy(float(kilometresDriven), float(litresUsed)) # Calling the function to calculate fuel economy
    print("Your fuel economy is: " + str(fuelEconomy) + " litres per 100km") # Print fuel economy
    print("would you like to add this data to a spreadsheet?")
    answer = input("Please enter yes or no: ") # User input to choose between yes or no
    if answer == "yes":
        insert_new_data(kilometresDriven, litresUsed)
elif type == "2": #summary data 
    book = load_workbook('testDocument.xlsx') # Load workbook
    sheet = book.active # Load active sheet
    dict = {}
    for row in sheet.iter_rows():
        if row[0].value not in dict.keys():
            temp = {str(row[0].value): [row[1].value, row[2].value]}
            dict.update(temp)
        else:
            dict[row[0].value].append("split")
            dict[row[0].value].append(row[1].value)
            dict[row[0].value].append(row[2].value)
    for key, value in dict.items():
        print(key, value)
        
    
    
