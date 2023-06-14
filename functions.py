from openpyxl import Workbook, load_workbook
from datetime import date
import xlsxwriter
import matplotlib.pyplot as plt
import os

#Current Date in DD/MM/YYYY format
today = date.today()
dateToday = today.strftime("%d/%m/%Y")
#today2 = "14/06/2023"

def calculate_fuel_economy(kilometresDriven, litresUsed): # Function to calculate fuel economy by litres per 100km
    fuelEconomy =  (litresUsed / kilometresDriven) * 100
    return fuelEconomy

def insert_new_data(date, distance, fuel, price): #Function to insert new data into spreadsheet
    try: # Try to open spreadsheet
        book = load_workbook('fuelEconomyData.xlsx') # Load workbook
        sheet = book.active # Load active sheet
    except: # If no spreadsheet -> create a new one 
        print("There is no appropriate excel file in the directory, creating a new one now")
        workbook = xlsxwriter.Workbook('fuelEconomyData.xlsx') # Create new workbook
        worksheet = workbook.add_worksheet() # Create new worksheet
        worksheet.write("A1", 'Date')# Add column headers
        worksheet.write("B1", 'Distance')
        worksheet.write("C1", 'Fuel')
        worksheet.write("D1", 'Price')
        workbook.close() # Close workbook
        book = load_workbook('fuelEconomyData.xlsx')
        sheet = book.active
    if date == "today":
        new_data = (dateToday, float(distance), float(fuel), float(price))
    else:
        new_data = (date, float(distance), float(fuel), float(price)) # Insert the new data into the spreadsheet
    sheet.append(new_data)
    book.save('fuelEconomyData.xlsx') # Save changes
    print("Data successfully added to spreadsheet")

def printFuelEconomy(fuelEconomy): # Print fuel economy to 2 decimal places 
    print("\nYour fuel economy is {:0.2f} litres per 100kms\n".format(fuelEconomy)) 

def returnFuelEconomy():
    kilometresDriven = input("Please enter the number of kilometres driven: ") # User input for kilometres driven
    litresUsed = input("Please enter the number of litres used: ") # User input for litres used
    price = input("Please enter the price of fuel: ") # User input for price of fuel
    dateChoice = input("do you want to use todays date? (yes or no): ") # User input to choose between yes or no
    if dateChoice == "no":
        finalDate = input("please enter the date in DD/MM/YYYY format: ")
    else:
        finalDate = "today"
    fuelEconomy = calculate_fuel_economy(float(kilometresDriven), float(litresUsed)) # Calling the function to calculate fuel economy
    printFuelEconomy(fuelEconomy)
    print("would you like to add this data to a spreadsheet?")
    answer = input("Please enter yes or no: ") # User input to choose between yes or no
    if answer == "yes":
        insert_new_data(finalDate, kilometresDriven, litresUsed, price)
    #elif answer == "no": -> ask if want to see summary data 
    
def generateSummaryData():
    os.system('cmd /c "streamlit run app.py" ')